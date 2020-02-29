import openpyxl

class Student:
    def __init__(self,fn="",ln="",ave=0):
        self.first_name=fn
        self.last_name=ln
        self.ave=ave
    def __le__(self,other):
        if self.ave<=other.ave:
            return True
        return False

def bubble (array = []):
    for i in range(len(array)-1):
        for j in range(len(array)-1):
            if array[j] <= array[j+1]:
                array[j],array[j+1] = array[j+1],array[j]
    return array
    

wb=openpyxl.load_workbook("students2.xlsx")
ws1=wb.active
ws2=wb.create_sheet("Sort Students")
ws2.cell(row=1,column=1,value="First Name")
ws2.cell(row=1,column=2,value="Last Name")
ws2.cell(row=1,column=3,value="Ave")

students=[]
for row in ws1.rows:
    if row[0].row == 1:
        continue
    s=Student(row[0].value,row[1].value,row[9].value)
    students.append(s)

bubble(students)

#count=2
for student in students:
    temp=[]
    temp.append(student.first_name)
    temp.append(student.last_name)
    temp.append(student.ave)
    ws2.append(temp)
#    ws2.cell(row=count,column=1,value=student.first_name)
#    ws2.cell(row=count,column=2,value=student.last_name)
#    ws2.cell(row=count,column=3,value=student.ave)
#    count+=1

wb.save("students2.xlsx")

wb.close
