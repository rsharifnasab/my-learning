import random
import openpyxl

wb=openpyxl.Workbook()
ws=wb.active
ws.title = "scores"

alphabet = "abcdefghijklmnopqrstuvwxyz";

ws.cell(row=1,column=1,value="First Name")
ws.cell(row=1,column=2,value="Last Name")
ws.cell(row=1,column=3,value="Score 1")
ws.cell(row=1,column=4,value="Score 2")
ws.cell(row=1,column=5,value="Score 3")
ws.cell(row=1,column=6,value="Score 4")
ws.cell(row=1,column=7,value="Score 5")

for i in range(2,52):
    name=""
    for j in range(5):
        name+=alphabet[random.randint(0,25)]
        if j==0:
            name = name.capitalize()
    ws.cell(row=i,column=1,value=name)
    
    name=""
    for j in range(5):
        name+=alphabet[random.randint(0,25)]
        if j==0:
            name=name.capitalize()
    ws.cell(row=i,column=2,value=name)
    
    for j in range(3,8):
        ws.cell(row=i,column=j,value=random.randint(50,100))

wb.save("students.xlsx")

wb.close
    
    