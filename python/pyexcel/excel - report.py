import openpyxl

wb=openpyxl.load_workbook("students2.xlsx")
ws1=wb["Sort Students"]
if "Report" in wb.sheetnames:
    ws2=wb["Report"]
else:
    ws2=wb.create_sheet("Report")

array = {'A':0,'B':0,'C':0}

for row in ws1.rows:
    if row[0].row == 1:
        continue
    if row[2].value>=70 and row[2].value<80:
        array['C']+=1
    elif row[2].value>=80 and row[2].value<90:
        array['B']+=1
    else:
        array['A']+=1

ws2.cell(row=1,column=1,value="Students in grade A : ")
ws2.cell(row=1,column=2,value=array['A'])
ws2.cell(row=2,column=1,value="Students in grade B : ")
ws2.cell(row=2,column=2,value=array['B'])
ws2.cell(row=3,column=1,value="Students in grade C : ")
ws2.cell(row=3,column=2,value=array['C'])

wb.save("students2.xlsx")

wb.close
