import openpyxl

wb=openpyxl.load_workbook("students.xlsx")

ws=wb.active

ws.cell(row=1,column=8,value="Min")
ws.cell(row=1,column=9,value="Max")
ws.cell(row=1,column=10,value="Ave")

count=2
for row in ws.iter_rows(min_row=1,max_row=50,min_col=1,max_col=5,row_offset=1,column_offset=2):
    Sum=0
    #Min=0
    #Max=0
    for cell in row:
        Sum+=cell.value
        if cell.column=='C':
            Min=cell.value
            Max=cell.value
        else:
            if Min>cell.value:
                Min=cell.value
            if Max<cell.value:
                Max=cell.value
    ws.cell(row=count,column=8,value=Min)
    ws.cell(row=count,column=9,value=Max)
    ws.cell(row=count,column=10,value=Sum/5)
    count+=1

wb.save("students2.xlsx")

wb.close